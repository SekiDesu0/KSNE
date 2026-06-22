from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash
from sqlalchemy import func, and_
from sqlalchemy.exc import IntegrityError
from datetime import date, datetime

from models.models import (
    db, Zona, Modulo, Producto, PrecioHistorico,
    Worker, Rendicion, RendicionItem,
)
from utils import (
    admin_required, validate_rut, format_rut, validate_phone,
    format_phone, generate_random_password, get_report_params,
)
from services import report_service, rendiciones_service


admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

BANCOS = [
    "Banco Estado", "Banco de Chile", "Banco Falabella", "BCI Nova",
    "Banco Bice", "Banco Consorcio", "Banco Security", "Banefe",
    "Coopeuch", "Corpbanca", "Multicaja", "Ahorrocoop",
    "BBVA Chile", "Banco Condell", "Banco Do Brasil S.A.",
    "Banco Edwards Citi", "Banco Internacional", "Banco Itaú Chile",
    "Banco París", "Banco Penta", "Banco Ripley",
    "Banco Santander Chile", "Banco de Crédito e Inversiones - BCI",
    "Banco de la Nación Argentina", "Caputal", "Coocretal",
    "Credichile Atlas", "DNB Bank ASA", "Detacoop",
    "Oriencoop", "RABOBANK CHILE", "Scotiabank Chile",
    "Servipag Express", "THE BANK OF TOKYO-MITSUBISHI UFJ, LTD.",
]


# ============================================================
# WORKERS
# ============================================================

@admin_bp.route('/workers', methods=['GET', 'POST'])
@admin_required
def manage_workers():
    form_data = {}

    if request.method == 'POST':
        raw_rut = request.form['rut']
        raw_phone = request.form['phone']
        name = request.form['name'].strip()
        modulo_id = request.form.get('modulo_id')
        tipo = request.form.get('tipo', 'Full Time')
        form_data = request.form

        if not validate_rut(raw_rut):
            flash("El RUT ingresado no es válido.", "danger")
        elif not validate_phone(raw_phone):
            flash("El teléfono debe tener 9 dígitos válidos.", "danger")
        elif not modulo_id:
            flash("Debes asignar un módulo al trabajador.", "danger")
        else:
            rut = format_rut(raw_rut)
            phone = format_phone(raw_phone)
            password = generate_random_password()
            p_hash = generate_password_hash(password)

            nombre_banco = request.form.get('nombre_banco', '')
            if nombre_banco == '__otro__':
                nombre_banco = request.form.get('nombre_banco_otro', '')

            new_worker = Worker(
                rut=rut, name=name, phone=phone, password_hash=p_hash,
                is_admin=False, modulo_id=int(modulo_id), tipo=tipo,
                nombre_banco=nombre_banco,
                numero_cuenta=request.form.get('numero_cuenta', ''),
                tipo_cuenta=request.form.get('tipo_cuenta', ''),
                rut_banco=request.form.get('rut_banco', ''),
            )
            try:
                db.session.add(new_worker)
                db.session.commit()
                flash(f"Trabajador guardado. Contraseña temporal: <strong>{password}</strong>", "success")
                return redirect(url_for('admin.manage_workers'))
            except IntegrityError:
                db.session.rollback()
                flash("El RUT ya existe en el sistema.", "danger")

    # Build (id, rut, name, phone, modulo_name, modulo_id, tipo) tuples
    # to preserve the existing template contract.
    workers_rows = (
        db.session.query(Worker, Modulo)
        .outerjoin(Modulo, Worker.modulo_id == Modulo.id)
        .filter(Worker.is_admin == False)
        .all()
    )
    workers = [
        (w.id, w.rut, w.name, w.phone, m.name if m else None, w.modulo_id, w.tipo,
         w.nombre_banco or '', w.numero_cuenta or '', w.tipo_cuenta or '', w.rut_banco or '')
        for w, m in workers_rows
    ]

    modulos_rows = (
        db.session.query(Modulo, Zona)
        .join(Zona, Modulo.zona_id == Zona.id)
        .order_by(Zona.name, Modulo.name)
        .all()
    )
    modulos = [(m.id, m.name, z.name) for m, z in modulos_rows]

    return render_template('admin_workers.html', workers=workers, form=form_data, modulos=modulos, bancos=BANCOS)


@admin_bp.route('/workers/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_worker(id):
    if request.method == 'POST':
        raw_phone = request.form['phone']
        name = request.form['name'].strip()
        modulo_id = request.form.get('modulo_id')
        tipo = request.form.get('tipo', 'Full Time')

        if not validate_phone(raw_phone):
            flash("El teléfono debe tener 9 dígitos válidos.", "danger")
            return redirect(url_for('admin.edit_worker', id=id))
        elif not modulo_id:
            flash("Debes seleccionar un módulo.", "danger")
            return redirect(url_for('admin.edit_worker', id=id))

        worker = db.session.get(Worker, id)
        if worker is None:
            return redirect(url_for('admin.manage_workers'))

        worker.name = name
        worker.phone = format_phone(raw_phone)
        worker.modulo_id = int(modulo_id)
        worker.tipo = tipo
        nombre_banco = request.form.get('nombre_banco', '')
        if nombre_banco == '__otro__':
            nombre_banco = request.form.get('nombre_banco_otro', '')
        worker.nombre_banco = nombre_banco
        worker.numero_cuenta = request.form.get('numero_cuenta', '')
        worker.tipo_cuenta = request.form.get('tipo_cuenta', '')
        worker.rut_banco = request.form.get('rut_banco', '')

        db.session.commit()
        flash("Trabajador actualizado exitosamente.", "success")
        return redirect(url_for('admin.manage_workers'))

    worker = db.session.get(Worker, id)

    modulos_rows = (
        db.session.query(Modulo, Zona)
        .join(Zona, Modulo.zona_id == Zona.id)
        .order_by(Zona.name, Modulo.name)
        .all()
    )
    modulos = [(m.id, m.name, z.name) for m, z in modulos_rows]

    if not worker:
        return redirect(url_for('admin.manage_workers'))

    worker_tuple = (worker.id, worker.rut, worker.name, worker.phone, worker.modulo_id,
                    worker.nombre_banco or '', worker.numero_cuenta or '', worker.tipo_cuenta or '',
                    worker.rut_banco or '')
    return render_template('edit_worker.html', worker=worker_tuple, modulos=modulos, bancos=BANCOS)


@admin_bp.route('/workers/delete/<int:id>', methods=['POST'])
@admin_required
def delete_worker(id):
    worker = db.session.get(Worker, id)
    if worker is not None:
        db.session.delete(worker)
        db.session.commit()
        flash("Trabajador eliminado.", "info")
    return redirect(url_for('admin.manage_workers'))


@admin_bp.route('/workers/reset_password/<int:id>', methods=['POST'])
@admin_required
def admin_reset_password(id):
    worker = db.session.get(Worker, id)
    if worker is None:
        return redirect(url_for('admin.manage_workers'))

    new_password = generate_random_password()
    worker.password_hash = generate_password_hash(new_password)
    db.session.commit()

    flash(f"Contraseña de {worker.name} restablecida. La nueva contraseña es: <strong>{new_password}</strong>", "warning")
    return redirect(url_for('admin.manage_workers'))


# ============================================================
# STRUCTURE (Zonas & Modulos)
# ============================================================

@admin_bp.route('/estructura', methods=['GET', 'POST'])
@admin_required
def manage_structure():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_zona':
            name = request.form.get('zona_name').strip()
            try:
                db.session.add(Zona(name=name))
                db.session.commit()
                flash("Zona guardada exitosamente.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("Ese nombre de Zona ya existe.", "danger")

        elif action == 'add_modulo':
            name = request.form.get('modulo_name').strip()
            zona_id = request.form.get('zona_id')
            if not zona_id:
                flash("Debes seleccionar una Zona válida.", "danger")
            else:
                db.session.add(Modulo(zona_id=int(zona_id), name=name))
                db.session.commit()
                flash("Módulo guardado exitosamente.", "success")

        return redirect(url_for('admin.manage_structure'))

    zonas = [(z.id, z.name) for z in Zona.query.order_by(Zona.name).all()]

    modulos_rows = (
        db.session.query(Modulo, Zona)
        .join(Zona, Modulo.zona_id == Zona.id)
        .order_by(Zona.name, Modulo.name)
        .all()
    )
    modulos = [(m.id, m.name, z.name) for m, z in modulos_rows]

    return render_template('admin_structure.html', zonas=zonas, modulos=modulos)


@admin_bp.route('/estructura/delete/<string:type>/<int:id>', methods=['POST'])
@admin_required
def delete_structure(type, id):
    try:
        if type == 'zona':
            count = db.session.query(func.count(Modulo.id)).filter(Modulo.zona_id == id).scalar()
            if count:
                flash("No puedes eliminar una Zona que tiene Módulos asignados.", "danger")
            else:
                zona = db.session.get(Zona, id)
                if zona is not None:
                    db.session.delete(zona)
                    flash("Zona eliminada.", "info")

        elif type == 'modulo':
            count = db.session.query(func.count(Worker.id)).filter(Worker.modulo_id == id).scalar()
            if count:
                flash("No puedes eliminar un Módulo que tiene Trabajadores asignados.", "danger")
            else:
                modulo = db.session.get(Modulo, id)
                if modulo is not None:
                    db.session.delete(modulo)
                    flash("Módulo eliminado.", "info")

        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Error al eliminar el registro.", "danger")

    return redirect(url_for('admin.manage_structure'))


# ============================================================
# PRODUCTS
# ============================================================

@admin_bp.route('/productos', methods=['GET', 'POST'])
@admin_required
def manage_products():
    if request.method == 'POST':
        name = request.form.get('name').strip()
        try:
            new_producto = Producto(name=name)
            db.session.add(new_producto)
            db.session.flush()  # Populate new_producto.id

            now = datetime.now()
            for zona in Zona.query.all():
                p_val = request.form.get(f'price_{zona.id}', '0')
                c_val = request.form.get(f'comm_{zona.id}', '0')
                price = int(str(p_val).replace('.', '').replace('$', '')) if p_val else 0
                commission = int(str(c_val).replace('.', '').replace('$', '')) if c_val else 0

                db.session.add(PrecioHistorico(
                    producto_id=new_producto.id,
                    zona_id=zona.id,
                    price=price,
                    commission=commission,
                    fecha_activacion=now,
                ))

            db.session.commit()
            flash("Producto maestro y sus precios creados exitosamente.", "success")
        except IntegrityError:
            db.session.rollback()
            flash("Ese producto ya existe en el catálogo.", "danger")

        return redirect(url_for('admin.manage_products'))

    zonas = [(z.id, z.name) for z in Zona.query.order_by(Zona.name).all()]

    # Find the most-recent vigente price for every (producto, zona) pair
    now = datetime.now()
    max_fecha_subq = (
        db.session.query(
            PrecioHistorico.producto_id.label('producto_id'),
            PrecioHistorico.zona_id.label('zona_id'),
            func.max(PrecioHistorico.fecha_activacion).label('max_fecha'),
        )
        .filter(PrecioHistorico.fecha_activacion <= now)
        .group_by(PrecioHistorico.producto_id, PrecioHistorico.zona_id)
        .subquery()
    )
    current_prices = (
        db.session.query(PrecioHistorico)
        .join(
            max_fecha_subq,
            and_(
                PrecioHistorico.producto_id == max_fecha_subq.c.producto_id,
                PrecioHistorico.zona_id == max_fecha_subq.c.zona_id,
                PrecioHistorico.fecha_activacion == max_fecha_subq.c.max_fecha,
            ),
        )
        .all()
    )
    price_map = {
        (p.producto_id, p.zona_id): (p.price, p.commission)
        for p in current_prices
    }

    productos = Producto.query.order_by(Producto.name).all()
    productos_dict = {}
    for p in productos:
        productos_dict[p.id] = {'id': p.id, 'name': p.name, 'precios': {}, 'futuros': []}
        for z in zonas:
            z_id, z_name = z
            price, comm = price_map.get((p.id, z_id), (None, None))
            productos_dict[p.id]['precios'][z_id] = {
                'zona_name': z_name,
                'price': price or 0,
                'commission': comm or 0,
            }

    # Scheduled future prices
    future_rows = (
        db.session.query(PrecioHistorico, Zona)
        .join(Zona, PrecioHistorico.zona_id == Zona.id)
        .filter(PrecioHistorico.fecha_activacion > now)
        .order_by(PrecioHistorico.fecha_activacion.asc())
        .all()
    )
    for ph, z in future_rows:
        if ph.producto_id in productos_dict:
            productos_dict[ph.producto_id]['futuros'].append({
                'id': ph.id,
                'zona_name': z.name,
                'price': ph.price,
                'commission': ph.commission,
                'fecha': ph.fecha_activacion,
            })

    return render_template('admin_productos.html', zonas=zonas, productos=productos_dict.values())


@admin_bp.route('/productos/delete/<int:id>', methods=['POST'])
@admin_required
def delete_product(id):
    try:
        PrecioHistorico.query.filter_by(producto_id=id).delete()
        producto = db.session.get(Producto, id)
        if producto is not None:
            db.session.delete(producto)
        db.session.commit()
        flash("Producto maestro y su historial eliminados.", "info")
    except IntegrityError:
        db.session.rollback()
        flash("No puedes eliminar este producto porque ya tiene ventas registradas. Cámbiale el precio a 0 en su lugar.", "danger")
    return redirect(url_for('admin.manage_products'))


@admin_bp.route('/productos/precios/<int:id>', methods=['POST'])
@admin_required
def update_product_prices(id):
    fecha_date = request.form.get('fecha_activacion_date')
    fecha_time = request.form.get('fecha_activacion_time') or '00:00'
    if fecha_date:
        fecha_activacion = datetime.strptime(f"{fecha_date} {fecha_time}", '%Y-%m-%d %H:%M').replace(second=0)
    else:
        fecha_activacion = datetime.now()

    for zona in Zona.query.all():
        z_id = str(zona.id)
        new_price = int(request.form.get(f'price_{z_id}', '0').replace('.', ''))
        new_comm = int(request.form.get(f'comm_{z_id}', '0').replace('.', ''))

        db.session.add(PrecioHistorico(
            producto_id=id,
            zona_id=zona.id,
            price=new_price,
            commission=new_comm,
            fecha_activacion=fecha_activacion,
        ))

    db.session.commit()
    flash(f"Precios actualizados. Entrarán en vigencia el {fecha_activacion}.", "success")
    return redirect(url_for('admin.manage_products'))


@admin_bp.route('/productos/precios/cancelar/<int:id>', methods=['POST'])
@admin_required
def cancel_scheduled_price(id):
    ph = db.session.get(PrecioHistorico, id)
    if ph is not None:
        db.session.delete(ph)
        db.session.commit()
        flash("Cambio de precio programado cancelado.", "info")
    return redirect(url_for('admin.manage_products'))


@admin_bp.route('/api/productos/<int:id>/historial')
@admin_required
def api_product_history(id):
    rows = (
        db.session.query(Zona.name, PrecioHistorico.price, PrecioHistorico.fecha_activacion)
        .join(PrecioHistorico, PrecioHistorico.zona_id == Zona.id)
        .filter(PrecioHistorico.producto_id == id)
        .order_by(PrecioHistorico.fecha_activacion.asc())
        .all()
    )
    history = []
    for zona_name, price, fecha in rows:
        if isinstance(fecha, datetime):
            fecha_str = fecha.strftime('%Y-%m-%d %H:%M:%S')
        else:
            fecha_str = str(fecha)
        history.append({'zona': zona_name, 'price': price, 'fecha': fecha_str})
    return jsonify(history)


# ============================================================
# RENDICIONES
# ============================================================

@admin_bp.route('/rendiciones')
@admin_required
def admin_rendiciones():
    hoy = date.today()
    fecha_inicio = request.args.get('fecha_inicio', f"{hoy.year}-{hoy.month:02d}-01")
    fecha_fin = request.args.get('fecha_fin', hoy.strftime('%Y-%m-%d'))
    zona_id_seleccionada = request.args.get('zona_id')
    modulo_id_seleccionado = request.args.get('modulo_id')

    rendiciones_completas = rendiciones_service.get_filtered_rendiciones(
        fecha_inicio, fecha_fin,
        zona_id_seleccionada, modulo_id_seleccionado,
    )
    workers, modulos, zonas, anios_disponibles = rendiciones_service.get_filter_catalogs()

    return render_template('admin_rendiciones.html',
                           rendiciones=rendiciones_completas,
                           workers=workers,
                           modulos=modulos,
                           zonas=zonas,
                           fecha_inicio=fecha_inicio,
                           fecha_fin=fecha_fin,
                           zona_actual=zona_id_seleccionada,
                           modulo_actual=modulo_id_seleccionado,
                           anios_disponibles=anios_disponibles)


@admin_bp.route('/rendiciones/delete/<int:id>', methods=['POST'])
@admin_required
def delete_rendicion(id):
    RendicionItem.query.filter_by(rendicion_id=id).delete()
    rendicion = db.session.get(Rendicion, id)
    if rendicion is not None:
        db.session.delete(rendicion)
    db.session.commit()
    flash("Rendición eliminada.", "info")
    return redirect(url_for('admin.admin_rendiciones'))


@admin_bp.route('/rendiciones/edit/<int:id>', methods=['POST'])
@admin_required
def edit_rendicion(id):
    fecha = request.form.get('fecha')
    worker_id = request.form.get('worker_id')
    modulo_id = request.form.get('modulo_id')
    companion_id = request.form.get('companion_id') or None
    if companion_id and worker_id == companion_id:
        flash("Error: No puedes asignarte a ti mismo como acompañante.", "danger")
        return redirect(url_for('admin.admin_rendiciones'))
    companion2_id = request.form.get('companion2_id') or None
    if companion2_id and worker_id == companion2_id:
        flash("Error: No puedes asignarte a ti mismo como acompañante 2.", "danger")
        return redirect(url_for('admin.admin_rendiciones'))
    worker_comision = 1 if request.form.get('worker_comision') else 0
    companion_comision = 1 if request.form.get('companion_comision') else 0
    companion2_comision = 1 if request.form.get('companion2_comision') else 0

    if not companion_id:
        companion_comision = 0
    if not companion2_id:
        companion2_comision = 0

    def clean_money(val):
        if not val:
            return 0
        return int(str(val).replace('.', '').replace('$', ''))

    try:
        debito = clean_money(request.form.get('venta_debito'))
        credito = clean_money(request.form.get('venta_credito'))
        mp = clean_money(request.form.get('venta_mp'))
        efectivo = clean_money(request.form.get('venta_efectivo'))

        bol_debito = int(request.form.get('boletas_debito') or 0)
        bol_credito = int(request.form.get('boletas_credito') or 0)
        bol_mp = int(request.form.get('boletas_mp') or 0)
        bol_efectivo = int(request.form.get('boletas_efectivo') or 0)

        gastos = clean_money(request.form.get('gastos'))
        observaciones = request.form.get('observaciones', '').strip()

        # 1. Update product quantities (rendicion_items)
        for key, value in request.form.items():
            if key.startswith('qty_'):
                ri_id = int(key.split('_')[1])
                nueva_qty = int(value or 0)
                item = db.session.get(RendicionItem, ri_id)
                if item is not None:
                    item.cantidad = nueva_qty

        # 2. Update the main rendicion row
        rendicion = db.session.get(Rendicion, id)
        if rendicion is not None:
            rendicion.fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            rendicion.worker_id = int(worker_id)
            rendicion.modulo_id = int(modulo_id)
            rendicion.companion_id = int(companion_id) if companion_id else None
            rendicion.companion2_id = int(companion2_id) if companion2_id else None
            rendicion.worker_comision = bool(worker_comision)
            rendicion.companion_comision = bool(companion_comision)
            rendicion.companion2_comision = bool(companion2_comision)
            rendicion.venta_debito = debito
            rendicion.venta_credito = credito
            rendicion.venta_mp = mp
            rendicion.venta_efectivo = efectivo
            rendicion.boletas_debito = bol_debito
            rendicion.boletas_credito = bol_credito
            rendicion.boletas_mp = bol_mp
            rendicion.boletas_efectivo = bol_efectivo
            rendicion.gastos = gastos
            rendicion.observaciones = observaciones

        db.session.commit()
        flash("Rendición y productos actualizados correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al actualizar: {str(e)}", "danger")

    return redirect(url_for('admin.admin_rendiciones'))


# ============================================================
# REPORTS
# ============================================================

@admin_bp.route('/reportes')
@admin_required
def admin_reportes_index():
    modulos_rows = (
        db.session.query(Modulo, Zona)
        .join(Zona, Modulo.zona_id == Zona.id)
        .order_by(Zona.name, Modulo.name)
        .all()
    )
    modulos = [(m.id, m.name, z.name) for m, z in modulos_rows]
    return render_template('admin_reportes_index.html', modulos=modulos)


@admin_bp.route('/reportes/modulo/<int:modulo_id>')
@admin_required
def report_modulo_periodo(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    mod_name, workers_list, anios_list = report_service.get_modulo_workers_and_anios(modulo_id)
    data = report_service.get_modulo_periodo_data(modulo_id, fecha_inicio, fecha_fin, worker_id)

    return render_template('admin_report_modulo.html',
                           modulo_name=mod_name, modulo_id=modulo_id,
                           mes_nombre=f"{fecha_inicio} a {fecha_fin}",
                           dias_en_periodo=data['dias_en_periodo'],
                           data_por_dia=data['data_por_dia'],
                           totales_mes=data['totales_mes'],
                           dias_activos=data['dias_activos'],
                           workers_list=workers_list, worker_actual=worker_id,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                           anios_disponibles=anios_list)


@admin_bp.route('/reportes/modulo/<int:modulo_id>/comisiones')
@admin_required
def report_modulo_comisiones(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    mod_name, workers_list, anios_list = report_service.get_modulo_workers_and_anios(modulo_id)
    data = report_service.get_comisiones_data(modulo_id, fecha_inicio, fecha_fin, worker_id)

    return render_template('admin_report_comisiones.html',
                           modulo_name=mod_name, modulo_id=modulo_id,
                           mes_nombre=f"{fecha_inicio} a {fecha_fin}",
                           workers_data=data['workers_data'],
                           dias_en_periodo=data['dias_en_periodo'],
                           workers_list=workers_list, worker_actual=worker_id,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                           anios_disponibles=anios_list)


@admin_bp.route('/reportes/modulo/<int:modulo_id>/horarios')
@admin_required
def report_modulo_horarios(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    mod_name, workers_list, anios_list = report_service.get_modulo_workers_and_anios(modulo_id)
    data = report_service.get_horarios_data(modulo_id, fecha_inicio, fecha_fin, worker_id)

    return render_template('admin_report_horarios.html',
                           modulo_name=mod_name, modulo_id=modulo_id,
                           mes_nombre=f"{fecha_inicio} a {fecha_fin}",
                           workers_data=data['workers_data'],
                           dias_en_periodo=data['dias_en_periodo'],
                           workers_list=workers_list, worker_actual=worker_id,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                           anios_disponibles=anios_list)


@admin_bp.route('/reportes/modulo/<int:modulo_id>/centros_comerciales')
@admin_required
def report_modulo_centros_comerciales(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    mod_name, workers_list, anios_list = report_service.get_modulo_workers_and_anios(modulo_id)
    data = report_service.get_cc_data(modulo_id, fecha_inicio, fecha_fin, worker_id)

    return render_template('admin_report_cc.html',
                           modulo_name=mod_name, modulo_id=modulo_id,
                           mes_nombre=f"{fecha_inicio} a {fecha_fin}",
                           dias_en_periodo=data['dias_en_periodo'],
                           data_por_dia=data['data_por_dia'],
                           totales=data['totales'],
                           workers_list=workers_list, worker_actual=worker_id,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                           anios_disponibles=anios_list)


@admin_bp.route('/reportes/modulo/<int:modulo_id>/calculo_iva')
@admin_required
def report_modulo_calculo_iva(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    mod_name, workers_list, anios_list = report_service.get_modulo_workers_and_anios(modulo_id)
    data = report_service.get_iva_data(modulo_id, fecha_inicio, fecha_fin, worker_id)

    return render_template('admin_report_iva.html',
                           modulo_name=mod_name, modulo_id=modulo_id,
                           mes_nombre=f"{fecha_inicio} a {fecha_fin}",
                           dias_en_periodo=data['dias_en_periodo'],
                           data_por_dia=data['data_por_dia'],
                           totales=data['totales'],
                           workers_list=workers_list, worker_actual=worker_id,
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                           anios_disponibles=anios_list)


@admin_bp.route('/reportes/modulo/<int:modulo_id>/exportar_excel')
@admin_required
def report_modulo_exportar_excel(modulo_id):
    fecha_inicio, fecha_fin, worker_id = get_report_params()
    data = report_service.get_modulo_periodo_data(modulo_id, fecha_inicio, fecha_fin, worker_id)
    mod_name, _, _ = report_service.get_modulo_workers_and_anios(modulo_id)

    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Ventas"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # ── column colors matching web table ──
    col_colors = {
        2: '198754',   # Venta Total → text-success green
        3: '0DCAF0',   # Comisión → text-info cyan
        4: 'DC3545',   # Gastos → text-danger red
        5: '6C757D',   # Crédito → text-muted gray
        6: '6C757D',   # Débito → text-muted gray
        7: '6C757D',   # Mercado Pago → text-muted gray
        8: '6C757D',   # Efectivo/Dep. → text-muted gray
        9: 'e5904d',   # Red. Crédito → custom orange
        10: 'e5904d',  # Red. Débito → custom orange
        11: 'e5904d',  # Red. MP → custom orange
        12: '20c997',  # REDELCOM Neto → teal
        13: 'FFC107',  # Efectivo - Gastos → text-warning
        14: '0D6EFD',  # Venta Total Neto → text-primary blue
    }

    hdr_fill = PatternFill(start_color="2B303A", end_color="2B303A", fill_type="solid")
    ws.merge_cells('A1:O1')
    ws['A1'] = f"Resumen Financiero — {mod_name} ({fecha_inicio} a {fecha_fin})"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Día', 'Venta Total', 'Comisión', 'Gastos',
               'Crédito', 'Débito', 'Mercado Pago', 'Efectivo/Dep.',
               'Red. Crédito', 'Red. Débito', 'Red. MP',
               'REDELCOM Neto', 'Efectivo - Gastos', 'Venta Total Neto']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        font_color = col_colors.get(col, 'FFFFFF')
        cell.font = Font(bold=True, color=font_color, size=11)

    for row_idx, dia in enumerate(data['dias_en_periodo'], 4):
        d = data['data_por_dia'][dia]
        vals = [
            dia,
            d['venta_total'],
            d['comision'],
            d['gastos'],
            d['credito'],
            d['debito'],
            d['mp'],
            d['efectivo'],
            d['credito'] * 0.97620,
            d['debito'] * 0.98453,
            d['mp'] * 0.98691,
            d['credito'] * 0.97620 + d['debito'] * 0.98453 + d['mp'] * 0.98691,
            d['efectivo'] - d['gastos'],
            d['credito'] * 0.97620 + d['debito'] * 0.98453 + d['mp'] * 0.98691 + d['efectivo'] - d['gastos'],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = border
            if col == 1:
                cell.alignment = center
            else:
                cell.number_format = '#,##0'
                if col in col_colors:
                    cell.font = Font(color=col_colors[col])

    total_row = 4 + len(data['dias_en_periodo'])
    totals = data['totales_mes']
    total_vals = [
        'TOTAL',
        totals['venta_total'],
        totals['comision'],
        totals['gastos'],
        totals['credito'],
        totals['debito'],
        totals['mp'],
        totals['efectivo'],
        totals['credito'] * 0.97620,
        totals['debito'] * 0.98453,
        totals['mp'] * 0.98691,
        totals['credito'] * 0.97620 + totals['debito'] * 0.98453 + totals['mp'] * 0.98691,
        totals['efectivo'] - totals['gastos'],
        totals['credito'] * 0.97620 + totals['debito'] * 0.98453 + totals['mp'] * 0.98691 + totals['efectivo'] - totals['gastos'],
    ]
    total_fill = PatternFill(start_color="2B303A", end_color="2B303A", fill_type="solid")
    for col, v in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row, column=col, value=v)
        cell.fill = total_fill
        cell.border = border
        if col == 1:
            cell.alignment = center
            cell.font = Font(bold=True, color="FFFFFF", size=11)
        else:
            cell.number_format = '#,##0'
            font_color = col_colors.get(col, 'FFFFFF')
            cell.font = Font(bold=True, color=font_color, size=11)

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 8

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_{mod_name}_{fecha_inicio}_{fecha_fin}.xlsx".replace(' ', '_')
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
